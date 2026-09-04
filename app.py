"""
GST Invoice Extractor & Analytics — Streamlit app
Zero-cost stack: Streamlit Community Cloud (hosting) + Google Gemini free tier (extraction)

SETUP:
1. Get a free Gemini API key: https://aistudio.google.com/app/apikey
2. Locally: create .streamlit/secrets.toml with:
       GEMINI_API_KEY = "your-key-here"
3. Run locally:  streamlit run app.py
4. Deploy free: push this repo to GitHub -> share.streamlit.io -> New app ->
   point at this file -> add GEMINI_API_KEY in the app's "Secrets" settings.
"""

import streamlit as st
import google.generativeai as genai
import json
import pandas as pd
from PIL import Image
from datetime import datetime
from decimal import Decimal, InvalidOperation
from concurrent.futures import ThreadPoolExecutor, as_completed
from fpdf import FPDF
import statistics
import time
import io
import time
import threading

st.set_page_config(page_title="Ledger — GST Invoice Intelligence", page_icon="🗒️", layout="wide")

# ============================================================================
# DESIGN SYSTEM — "The Ledger"
# Token system:
#   Color   ink #10151C (bg) / surface #1A222C / paper #F7F2E7 (inset tables only)
#           brass #C9A468 (primary accent) / verified #3F8F5F / flagged #C1553A
#           text #EDE7D9 (on ink) / muted #93A0AC
#   Type    Display: Fraunces (editorial serif, ledger-masthead character)
#           Body: IBM Plex Sans   Data/numbers: IBM Plex Mono (tabular, aligns decimals)
#   Layout  Dark ink dashboard with paper-inset panels for tables — evokes a real
#           ledger book under a desk lamp, not a generic SaaS admin theme.
#   Signature: rotated ink-stamp badge for Verified/Needs Review, echoing the
#           "Authorised Signatory" rubber stamps on the real invoices this reads.
# ============================================================================
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,400;9..144,600;9..144,700&family=IBM+Plex+Sans:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;500;600&display=swap');

:root {
    --ink: #F7F4EC;
    --surface: #FFFFFF;
    --surface-2: #F1ECDF;
    --paper: #FFFFFF;
    --paper-line: #DCD3BD;
    --brass: #A97C34;
    --brass-dim: #8A6F49;
    --verified: #2C7A4B;
    --flagged: #B23F26;
    --text: #211A10;
    --muted: #5C5546;
}

/* base canvas */
.stApp {
    background: radial-gradient(ellipse 120% 80% at 50% -10%, #FFFFFF 0%, var(--ink) 55%) !important;
    color: var(--text);
}
[data-testid="stSidebar"] {
    background: var(--surface) !important;
    border-right: 1px solid rgba(201,164,104,0.18);
}
[data-testid="stHeader"] { background: transparent !important; }

/* typography */
h1, h2, h3 { font-family: 'Fraunces', serif !important; color: var(--text) !important; letter-spacing: -0.01em; }
h1 { font-weight: 600 !important; }
h1::after {
    content: ""; display: block; width: 64px; height: 2px; margin-top: 14px;
    background: linear-gradient(90deg, var(--brass), transparent);
}
body, p, div, span, label, .stMarkdown { font-family: 'IBM Plex Sans', sans-serif; }
.stCaption, [data-testid="stCaptionContainer"] { color: var(--muted) !important; font-family: 'IBM Plex Sans', sans-serif; }

/* every number in this app renders in tabular mono — decimals align like a ledger */
[data-testid="stMetricValue"], [data-testid="stDataFrame"], .stDataFrame,
[data-testid="stMarkdownContainer"] code {
    font-family: 'IBM Plex Mono', monospace !important;
}

/* metric cards -> brass-edged tiles */
[data-testid="stMetric"] {
    background: var(--surface-2);
    border: 1px solid rgba(201,164,104,0.25);
    border-left: 3px solid var(--brass);
    border-radius: 6px;
    padding: 16px 18px 12px 18px;
}
[data-testid="stMetricLabel"] { color: var(--muted) !important; font-family: 'IBM Plex Sans'; font-size: 0.8rem !important; text-transform: uppercase; letter-spacing: 0.05em; }
[data-testid="stMetricValue"] { color: var(--brass) !important; font-weight: 600 !important; }

/* tabs -> ledger index tabs */
.stTabs [data-baseweb="tab-list"] { gap: 4px; border-bottom: 1px solid rgba(201,164,104,0.25); }
.stTabs [data-baseweb="tab"] {
    background: transparent; color: var(--muted); font-family: 'IBM Plex Sans'; font-weight: 500;
    border-radius: 6px 6px 0 0; padding: 10px 18px;
}
.stTabs [aria-selected="true"] { color: var(--brass) !important; background: var(--surface-2) !important; border-bottom: 2px solid var(--brass); }

/* buttons */
.stButton button, .stFormSubmitButton button, .stDownloadButton button {
    font-family: 'IBM Plex Sans'; font-weight: 600; border-radius: 5px;
    border: 1px solid var(--brass-dim);
    background: linear-gradient(180deg, #C9A468 0%, #B08D57 100%);
    color: #14100A !important;
    transition: all 0.15s ease;
}
.stButton button:hover, .stFormSubmitButton button:hover, .stDownloadButton button:hover {
    box-shadow: 0 0 0 3px rgba(201,164,104,0.25); border-color: var(--brass);
}

/* dataframes / tables -> inset paper panel, like a bound ledger page */
[data-testid="stDataFrame"] {
    background: var(--paper) !important; border-radius: 4px; padding: 4px;
    border: 1px solid var(--paper-line);
}

/* expanders -> case files */
.streamlit-expanderHeader, [data-testid="stExpander"] {
    background: var(--surface-2) !important; border: 1px solid rgba(201,164,104,0.2) !important; border-radius: 6px !important;
}

/* alerts recolored to the ledger palette instead of Streamlit's default blue/orange/green */
div[data-testid="stAlert"], div[data-baseweb="notification"], .stAlert {
    border-radius: 6px !important;
    border: 1px solid rgba(201,164,104,0.3) !important;
    background: var(--surface-2) !important;
}
div[data-testid="stAlert"] p, div[data-testid="stAlert"] span,
div[data-testid="stAlert"] [data-testid="stMarkdownContainer"] {
    color: var(--text) !important;
    font-family: 'IBM Plex Sans', sans-serif !important;
}
div[data-testid="stAlert"] svg { fill: var(--brass) !important; }
/* warning-toned alerts lean rust, success-toned lean verified-green, matching the stamp badges */
div[data-testid="stAlert"]:has(svg[data-icon="warning"]) { border-left: 3px solid var(--flagged) !important; }
div[data-testid="stAlert"]:has(svg[data-icon="check"]) { border-left: 3px solid var(--verified) !important; }
div[data-testid="stAlert"]:has(svg[data-icon="info"]) { border-left: 3px solid var(--brass) !important; }

/* text inputs / selects */
.stTextInput input, .stSelectbox div[data-baseweb="select"], .stTextArea textarea {
    background: var(--surface-2) !important; color: var(--text) !important;
    border: 1px solid rgba(201,164,104,0.25) !important; border-radius: 5px !important;
}

/* the signature element: a rotated ink-stamp badge for Verified / Needs Review */
.ledger-stamp {
    display: inline-flex; align-items: center; gap: 6px;
    font-family: 'IBM Plex Mono', monospace; font-weight: 600; font-size: 0.78rem;
    letter-spacing: 0.06em; text-transform: uppercase;
    padding: 4px 12px; border-radius: 999px;
    transform: rotate(-2deg);
}
.ledger-stamp.verified { border: 1.5px solid var(--verified); color: var(--verified); background: rgba(63,143,95,0.08); }
.ledger-stamp.flagged { border: 1.5px solid var(--flagged); color: var(--flagged); background: rgba(193,85,58,0.08); }
.ledger-stamp.admin { border: 1.5px solid var(--brass); color: var(--brass); background: rgba(201,164,104,0.1); }

/* masthead divider under the title */
.ledger-masthead {
    font-family: 'IBM Plex Mono', monospace; color: var(--muted); font-size: 0.8rem;
    letter-spacing: 0.08em; text-transform: uppercase; margin-top: -8px; margin-bottom: 18px;
}

/* ============================================================================
   FUTURISTIC LAYER — glassmorphism, glow, 3D depth
   Adds: animated mesh backdrop, glass panels with real depth (blur + inner glow),
   numbered section eyebrows to break the page into distinct zones, and a tilt-on-hover
   3D interaction for cards, on top of the existing ledger palette/typography.
   ============================================================================ */

@keyframes meshDrift {
    0%   { transform: translate(0%, 0%) rotate(0deg); }
    50%  { transform: translate(-4%, 3%) rotate(6deg); }
    100% { transform: translate(0%, 0%) rotate(0deg); }
}
.stApp::before {
    content: ""; position: fixed; inset: -20%; z-index: 0; pointer-events: none;
    background:
        radial-gradient(circle at 15% 20%, rgba(201,164,104,0.10) 0%, transparent 35%),
        radial-gradient(circle at 85% 15%, rgba(63,143,214,0.08) 0%, transparent 40%),
        radial-gradient(circle at 50% 90%, rgba(63,143,95,0.06) 0%, transparent 45%);
    animation: meshDrift 22s ease-in-out infinite;
}
[data-testid="stAppViewContainer"] { position: relative; z-index: 1; }

/* glass panel — the base surface for section content */
.glass-panel {
    background: linear-gradient(180deg, rgba(255,255,255,0.80) 0%, rgba(241,236,223,0.70) 100%);
    backdrop-filter: blur(18px) saturate(140%);
    -webkit-backdrop-filter: blur(18px) saturate(140%);
    border: 1px solid rgba(169,124,52,0.22);
    border-radius: 14px;
    box-shadow: 0 8px 32px rgba(0,0,0,0.08), inset 0 1px 0 rgba(255,255,255,0.6);
    padding: 22px 26px;
    margin-bottom: 4px;
}

/* section eyebrow — numbered zone marker, gives the page distinct "sections" */
.section-eyebrow {
    display: flex; align-items: center; gap: 12px; margin: 38px 0 6px 0;
}
.section-eyebrow .num {
    font-family: 'IBM Plex Mono', monospace; font-size: 0.75rem; font-weight: 600;
    color: #2A1D0E; background: linear-gradient(135deg, var(--brass), #E0C48C);
    padding: 3px 9px; border-radius: 4px; letter-spacing: 0.05em;
    box-shadow: 0 0 16px rgba(201,164,104,0.45);
}
.section-eyebrow .label {
    font-family: 'IBM Plex Mono', monospace; font-size: 0.78rem; color: var(--muted);
    letter-spacing: 0.12em; text-transform: uppercase;
}
.section-eyebrow .rule {
    flex: 1; height: 1px;
    background: linear-gradient(90deg, rgba(201,164,104,0.4), transparent);
}

/* metric tiles get real 3D depth + hover tilt */
[data-testid="stMetric"] {
    transform-style: preserve-3d; perspective: 800px;
    transition: transform 0.25s ease, box-shadow 0.25s ease;
    background: linear-gradient(155deg, rgba(255,255,255,0.90), rgba(241,236,223,0.75)) !important;
    backdrop-filter: blur(12px);
    box-shadow: 0 6px 20px rgba(0,0,0,0.08), inset 0 1px 0 rgba(255,255,255,0.6);
}
[data-testid="stMetric"]:hover {
    transform: translateY(-3px) rotateX(4deg);
    box-shadow: 0 14px 32px rgba(169,124,52,0.18), inset 0 1px 0 rgba(255,255,255,0.6);
    border-color: rgba(169,124,52,0.5) !important;
}

/* expanders as glass cards with hover glow */
[data-testid="stExpander"] {
    background: linear-gradient(155deg, rgba(255,255,255,0.85), rgba(241,236,223,0.7)) !important;
    backdrop-filter: blur(10px);
    transition: box-shadow 0.2s ease, border-color 0.2s ease;
}
[data-testid="stExpander"]:hover { border-color: rgba(169,124,52,0.45) !important; box-shadow: 0 0 24px rgba(169,124,52,0.08); }

/* the 3D holographic cube in the hero */
.hero-3d-wrap {
    display: flex; align-items: center; justify-content: center; height: 200px;
    perspective: 1000px;
}
.hero-cube {
    position: relative; width: 84px; height: 84px; transform-style: preserve-3d;
    animation: cubeSpin 14s linear infinite;
}
.hero-cube .face {
    position: absolute; width: 84px; height: 84px;
    border: 1px solid rgba(201,164,104,0.65);
    background: linear-gradient(135deg, rgba(201,164,104,0.14), rgba(63,143,214,0.08));
    box-shadow: inset 0 0 24px rgba(201,164,104,0.15);
}
.hero-cube .front  { transform: translateZ(42px); }
.hero-cube .back   { transform: rotateY(180deg) translateZ(42px); }
.hero-cube .right  { transform: rotateY(90deg) translateZ(42px); }
.hero-cube .left   { transform: rotateY(-90deg) translateZ(42px); }
.hero-cube .top    { transform: rotateX(90deg) translateZ(42px); }
.hero-cube .bottom { transform: rotateX(-90deg) translateZ(42px); }
@keyframes cubeSpin {
    0%   { transform: rotateX(0deg) rotateY(0deg); }
    100% { transform: rotateX(360deg) rotateY(360deg); }
}

@keyframes scanline {
    0% { transform: translateY(-100%); opacity: 0; }
    10% { opacity: 0.5; }
    90% { opacity: 0.5; }
    100% { transform: translateY(2100%); opacity: 0; }
}
.hero-scan {
    position: absolute; left: 0; right: 0; height: 2px;
    background: linear-gradient(90deg, transparent, var(--brass), transparent);
    animation: scanline 5s linear infinite;
}

/* ============================================================================
   REFINEMENT PASS — frame the page like a designed product, not a default
   Streamlit stack: constrained content width, animated gradient title, glass
   forms/dropzones, custom scrollbar, and motion on every interactive control.
   ============================================================================ */

.block-container {
    max-width: 1180px; margin: 0 auto; padding-top: 2.2rem !important; padding-bottom: 4rem;
}

@keyframes titleSweep {
    0%   { background-position: 0% 50%; }
    50%  { background-position: 100% 50%; }
    100% { background-position: 0% 50%; }
}
h1 {
    background: linear-gradient(100deg, var(--text) 20%, var(--brass) 45%, #E0C48C 55%, var(--text) 80%);
    background-size: 220% auto;
    -webkit-background-clip: text; background-clip: text; -webkit-text-fill-color: transparent;
    animation: titleSweep 8s ease-in-out infinite;
}
h1::after { background: linear-gradient(90deg, var(--brass), transparent) !important; -webkit-text-fill-color: initial; }

[data-testid="stFileUploaderDropzone"] {
    background: linear-gradient(155deg, rgba(255,255,255,0.85), rgba(241,236,223,0.6)) !important;
    border: 1.5px dashed rgba(169,124,52,0.45) !important;
    border-radius: 12px !important;
    backdrop-filter: blur(10px);
    transition: border-color 0.2s ease, box-shadow 0.2s ease;
}
[data-testid="stFileUploaderDropzone"]:hover { border-color: var(--brass) !important; box-shadow: 0 0 28px rgba(169,124,52,0.15); }
[data-testid="stForm"] {
    background: linear-gradient(155deg, rgba(255,255,255,0.80), rgba(241,236,223,0.65)) !important;
    backdrop-filter: blur(14px); border: 1px solid rgba(169,124,52,0.2) !important;
    border-radius: 14px !important; padding: 22px 24px !important;
}
[data-testid="stDataFrameResizable"], .stDataEditor { border-radius: 8px !important; overflow: hidden; }
.stTextInput input:focus, .stTextArea textarea:focus, .stSelectbox div[data-baseweb="select"]:focus-within {
    border-color: var(--brass) !important; box-shadow: 0 0 0 3px rgba(201,164,104,0.18) !important; outline: none !important;
}
div[data-baseweb="popover"] ul, div[data-baseweb="menu"] { background: var(--surface-2) !important; border: 1px solid rgba(201,164,104,0.25) !important; }
div[data-baseweb="popover"] li { color: var(--text) !important; }
div[data-baseweb="popover"] li:hover { background: rgba(201,164,104,0.15) !important; }
[data-testid="stProgress"] > div > div > div { background: linear-gradient(90deg, var(--brass), #E0C48C) !important; }
::-webkit-scrollbar { width: 10px; height: 10px; }
::-webkit-scrollbar-track { background: var(--ink); }
::-webkit-scrollbar-thumb { background: rgba(201,164,104,0.35); border-radius: 6px; }
::-webkit-scrollbar-thumb:hover { background: rgba(201,164,104,0.55); }
@keyframes fadeUp { from { opacity: 0; transform: translateY(14px); } to { opacity: 1; transform: translateY(0); } }
.glass-panel { animation: fadeUp 0.6s ease both; }
.stTabs [aria-selected="true"] { box-shadow: 0 2px 12px rgba(201,164,104,0.25); }
.sidebar-divider { height: 1px; margin: 18px 0; background: linear-gradient(90deg, transparent, rgba(201,164,104,0.4), transparent); }

/* premium workspace surfaces */
[data-testid="stSidebar"] > div:first-child { padding: 1.4rem 1.1rem; }
[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p { color: #5C5546; line-height: 1.55; }
.hero-shell {
    position: relative; overflow: hidden; min-height: 248px; margin: 0 0 26px 0;
    display: grid; grid-template-columns: 1.12fr .88fr; align-items: center; gap: 18px;
    padding: 28px 34px; border-radius: 24px;
    background: radial-gradient(circle at 86% 50%, rgba(169,124,52,.18), transparent 32%), radial-gradient(circle at 60% 10%, rgba(79,137,196,.10), transparent 34%), linear-gradient(135deg, #F3ECDC 0%, #EFE6D0 100%);
    border: 1px solid rgba(169,124,52,.30);
    box-shadow: 0 22px 70px rgba(0,0,0,.10), inset 0 1px 0 rgba(255,255,255,.6);
}
.hero-shell:after { content:""; position:absolute; inset:0; pointer-events:none; background: linear-gradient(115deg, transparent 40%, rgba(255,255,255,.05) 50%, transparent 60%); transform: translateX(-75%); animation: sheen 9s ease-in-out infinite; }
@keyframes sheen { 0%,55% { transform: translateX(-75%); } 75%,100% { transform: translateX(75%); } }
.hero-copy { position: relative; z-index: 2; }
.hero-kicker { color: #8F6A2C; font: 600 .72rem/1.2 'IBM Plex Mono', monospace; letter-spacing: .17em; text-transform: uppercase; margin-bottom: 12px; }
.hero-title { margin: 0; color: #201A10; font: 600 clamp(2.2rem, 5vw, 4.2rem)/.95 'Fraunces', serif; letter-spacing: -.045em; }
.hero-subtitle { max-width: 520px; color: #5C5546; font-size: 1rem; line-height: 1.55; margin: 16px 0 0; }
.hero-pills { display:flex; flex-wrap:wrap; gap:8px; margin-top:20px; }
.hero-pill { padding:7px 11px; border-radius:999px; border:1px solid rgba(169,124,52,.30); background:rgba(255,255,255,.55); color:#3A3226; font:500 .7rem 'IBM Plex Mono', monospace; }
.hero-visual { position:relative; z-index:1; min-height:208px; display:flex; align-items:center; justify-content:center; }
.hero-orbit, .hero-orbit:before, .hero-orbit:after { position:absolute; content:""; border:1px solid rgba(201,164,104,.30); border-radius:50%; transform:rotateX(64deg) rotateZ(18deg); }
.hero-orbit { width:210px; height:92px; box-shadow:0 0 32px rgba(201,164,104,.12); animation: orbitFloat 6s ease-in-out infinite; }
.hero-orbit:before { inset:18px -18px; transform:rotateX(68deg) rotateZ(-32deg); border-color:rgba(92,159,213,.24); }
.hero-orbit:after { inset:-18px 24px; transform:rotateX(68deg) rotateZ(72deg); border-color:rgba(63,143,95,.22); }
@keyframes orbitFloat { 0%,100% { transform: translateY(5px) rotateX(64deg) rotateZ(18deg); } 50% { transform: translateY(-8px) rotateX(64deg) rotateZ(18deg); } }
.hero-cube { filter: drop-shadow(0 0 22px rgba(201,164,104,.28)); }
.workflow-strip { display:flex; gap:10px; flex-wrap:wrap; margin:0 0 20px; padding:12px 14px; border:1px solid rgba(169,124,52,.20); border-radius:14px; background:rgba(255,255,255,.55); }
.workflow-step { display:flex; gap:8px; align-items:center; color:#4A4438; font:500 .73rem 'IBM Plex Mono', monospace; }
.workflow-step b { display:grid; place-items:center; width:24px; height:24px; border-radius:8px; color:#19140C; background:linear-gradient(145deg,#E4C68B,#B9925D); font-size:.68rem; }
.workflow-arrow { color:#8C8370; }
[data-testid="stFileUploader"] { padding: 8px; border: 1px dashed rgba(169,124,52,.35); border-radius: 16px; background: linear-gradient(145deg, rgba(255,255,255,.6), rgba(255,255,255,.3)); }
[data-testid="stDataFrame"] { box-shadow: 0 14px 36px rgba(0,0,0,.08); }
@media (max-width: 800px) { .hero-shell { grid-template-columns: 1fr; padding: 24px; } .hero-visual { min-height: 160px; } .hero-title { font-size: 2.7rem; } }
</style>
""", unsafe_allow_html=True)


def stamp(label, kind="verified"):
    """Renders the ink-stamp badge — the app's signature visual element."""
    return f'<span class="ledger-stamp {kind}">● {label}</span>'


def section(num, label):
    """Numbered section eyebrow — breaks the page into distinct visual zones."""
    st.markdown(
        f'<div class="section-eyebrow"><span class="num">{num}</span>'
        f'<span class="label">{label}</span><span class="rule"></span></div>',
        unsafe_allow_html=True
    )


def render_hero():
    """Premium hero surface with CSS-only 3D orbital geometry for fast Streamlit loading."""
    st.markdown("""
        <section class="hero-shell">
            <div class="hero-copy">
                <div class="hero-kicker">GST intelligence / workspace 01</div>
                <h1 class="hero-title">Read the invoice.<br><em>Know the ledger.</em></h1>
                <p class="hero-subtitle">Turn handwritten and printed GST invoices into a verified, searchable financial picture — exact to the rupee.</p>
                <div class="hero-pills"><span class="hero-pill">AI vision extraction</span><span class="hero-pill">Exact math checks</span><span class="hero-pill">Export-ready</span></div>
            </div>
            <div class="hero-visual">
                <div class="hero-orbit"></div>
                <div class="hero-3d-wrap">
                    <div class="hero-cube">
                        <div class="face front"></div><div class="face back"></div><div class="face right"></div>
                        <div class="face left"></div><div class="face top"></div><div class="face bottom"></div>
                    </div>
                </div>
            </div>
            <div class="hero-scan"></div>
        </section>
    """, unsafe_allow_html=True)

# ---------- CONFIG: multi-key pool with automatic fallback ----------
# Reads GEMINI_API_KEY, GEMINI_API_KEY_2, GEMINI_API_KEY_3, ... from secrets.
# When a key hits its rate limit (5/min or 20/day on free tier), the app rotates
# to the next key automatically instead of failing. Each key needs its own free
# Google account + its own key at aistudio.google.com/app/apikey.
API_KEYS = []
primary = st.secrets.get("GEMINI_API_KEY", "")
if primary:
    API_KEYS.append(primary)
i = 2
while True:
    k = st.secrets.get(f"GEMINI_API_KEY_{i}", "")
    if not k:
        break
    API_KEYS.append(k)
    i += 1

if not API_KEYS:
    st.error("No GEMINI_API_KEY found in secrets. Add it in .streamlit/secrets.toml (local) "
              "or your Streamlit Cloud app's Secrets settings.")
    st.stop()



def get_model():
    """Return a model configured with the currently active key and stable free-tier model."""
    idx = get_active_key_index() % len(API_KEYS)
    genai.configure(api_key=API_KEYS[idx])
    model_name = st.secrets.get("GEMINI_MODEL", "gemini-3.6-flash")
    return genai.GenerativeModel(model_name)


def rotate_to_next_key():
    """Advance to the next key; return False when all configured keys are exhausted."""
    next_idx = get_active_key_index() + 1
    set_active_key_index(next_idx)
    return next_idx < len(API_KEYS)


EXTRACTION_PROMPT = """You are reading a handwritten or printed Indian GST tax invoice photo.
Extract every field into STRICT JSON only — no markdown fences, no commentary, no rounding.

Schema:
{
  "vendor_name": "", "vendor_gstin": "", "bill_no": "", "date": "",
  "customer_name": "", "customer_gstin": "",
  "line_items": [
    {"particulars": "", "hsn_code": "", "qty": "", "rate": "", "amount": ""}
  ],
  "subtotal": "",
  "cgst_pct": "", "cgst_amt": "",
  "sgst_pct": "", "sgst_amt": "",
  "igst_pct": "", "igst_amt": "",
  "grand_total": "",
  "uncertain_money_fields": ["ONLY numeric/money fields you weren't sure of, by name/index — e.g. 'line_items[2].amount', 'grand_total'. Do NOT include text fields like customer_name, customer_gstin, vendor_name, date, or particulars here — those don't affect GST accuracy."],
  "extraction_confidence": "high, medium, or low based on handwriting legibility of the MONEY figures specifically"
}

HARD RULES — these are non-negotiable:
1. Copy every number EXACTLY as written, digit for digit, as a string (e.g. "1500.00", not 1500).
   Do not round. Do not convert. Do not "clean up" a number to make totals match.
2. NEVER guess a digit you cannot clearly see in a MONEY field (qty, rate, amount, subtotal, tax
   amounts, grand total). If a money digit is illegible or ambiguous, put your best-guess value in
   the field AND add that field to "uncertain_money_fields".
3. Text fields (customer name, customer GSTIN, vendor name, date, particulars/product names) are
   often genuinely blank or messy on real invoices — that is normal and not something to flag.
   If blank, use "". If legible even loosely, just transcribe it — do not add these to
   uncertain_money_fields under any circumstances, since they don't affect GST calculation accuracy.
4. Do NOT attempt to fix or reconcile math yourself. Copy numbers exactly as written even if they
   don't seem to add up — a separate process checks the math. Your only job is faithful
   transcription of what's on the page, not correction.
5. If a numeric field is genuinely not present on the invoice (e.g. no IGST line at all), use ""
   (not 0 — 0 implies you saw a zero, "" means absent).
"""


def to_decimal(val):
    """Convert an extracted string field to Decimal without any float rounding.
    Returns (Decimal or None, was_parseable: bool)."""
    if val is None or val == "":
        return None, True  # legitimately absent, not an error
    s = str(val).strip().replace(",", "").replace("₹", "").replace("Rs.", "").replace("Rs", "")
    try:
        return Decimal(s), True
    except InvalidOperation:
        return None, False


def extract_leading_number(val):
    """For fields like qty that sometimes carry a unit (e.g. '3mtr', '15 nos'),
    pull out the leading numeric portion for math purposes. Returns (Decimal or None, ok)."""
    if val is None or val == "":
        return None, True
    s = str(val).strip().replace(",", "")
    num_chars = []
    for ch in s:
        if ch.isdigit() or ch == "." or (ch == "-" and not num_chars):
            num_chars.append(ch)
        else:
            break
    if not num_chars:
        return None, False
    try:
        return Decimal("".join(num_chars)), True
    except InvalidOperation:
        return None, False


def verify_invoice_math(inv):
    """Recompute everything with exact Decimal arithmetic. Never overwrite the
    extracted numbers — only flag discrepancies so a human can check the photo."""
    issues = []
    line_total = Decimal("0")

    for idx, item in enumerate(inv.get("line_items", [])):
        qty, qty_ok = extract_leading_number(item.get("qty"))       # tolerates "3mtr", "15 nos" etc.
        rate, rate_ok = to_decimal(item.get("rate"))
        amount, amt_ok = to_decimal(item.get("amount"))

        # Amount always counts toward the subtotal check if it parsed, regardless of
        # whether qty/rate parsed — a unit like "mtr" shouldn't hide a real amount.
        if amt_ok and amount is not None:
            line_total += amount

        if not amt_ok:
            issues.append(f"Line {idx+1}: amount field unparseable, needs manual check.")
            continue
        if not (qty_ok and rate_ok):
            continue  # qty/rate had non-numeric units we can't cross-check, but amount still counted above

        if qty is not None and rate is not None and amount is not None:
            expected = (qty * rate).quantize(Decimal("0.01"))
            actual = amount.quantize(Decimal("0.01"))
            if expected != actual:
                issues.append(
                    f"Line {idx+1} ({item.get('particulars','')}): qty×rate = {expected}, "
                    f"but invoice shows amount = {actual}. Diff = {actual - expected}."
                )

    subtotal, sub_ok = to_decimal(inv.get("subtotal"))
    cgst, _ = to_decimal(inv.get("cgst_amt"))
    sgst, _ = to_decimal(inv.get("sgst_amt"))
    igst, _ = to_decimal(inv.get("igst_amt"))
    grand_total, gt_ok = to_decimal(inv.get("grand_total"))

    if sub_ok and subtotal is not None and line_total != Decimal("0"):
        if subtotal.quantize(Decimal("0.01")) != line_total.quantize(Decimal("0.01")):
            issues.append(
                f"Sum of line-item amounts = {line_total.quantize(Decimal('0.01'))}, "
                f"but invoice subtotal shows {subtotal.quantize(Decimal('0.01'))}."
            )

    if gt_ok and grand_total is not None and sub_ok and subtotal is not None:
        tax_sum = (cgst or Decimal("0")) + (sgst or Decimal("0")) + (igst or Decimal("0"))
        expected_gt = (subtotal + tax_sum).quantize(Decimal("0.01"))
        actual_gt = grand_total.quantize(Decimal("0.01"))
        if expected_gt != actual_gt:
            issues.append(
                f"Subtotal + tax = {expected_gt}, but invoice grand total shows {actual_gt}. "
                f"Diff = {actual_gt - expected_gt}."
            )

    return issues


def d(val, default="0"):
    """Safe Decimal accessor for display/aggregation — never silently rounds a real number,
    only substitutes 0 when the field was genuinely blank."""
    dec, ok = to_decimal(val)
    return dec if (ok and dec is not None) else Decimal(default)


def prep_image_for_upload(pil_img, max_dimension=1600):
    """Downscale + normalize before sending to the API. Phone photos are often
    3000-4000px / several MB — the model reads handwriting just as accurately at
    ~1600px on the longest side, but the request transmits and processes far faster.
    This is the single biggest speed lever available without changing accuracy."""
    img = pil_img.convert("RGB")
    w, h = img.size
    if max(w, h) > max_dimension:
        scale = max_dimension / max(w, h)
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
    return img


class RateLimitError(Exception):
    """Raised only when EVERY key in the pool is exhausted, so the UI can show a
    clear, specific message instead of a generic parse error."""
    pass


def _is_rate_limit_error(e):
    msg = str(e)
    return "429" in msg or "quota" in msg.lower() or "rate limit" in msg.lower() or "ResourceExhausted" in msg


def extract_one(uf_name, pil_img, customer_tag, max_retries_per_key=2):
    """Runs one extraction call. On a rate limit, retries briefly with backoff (helps
    with the per-minute cap), and if that key is truly exhausted (daily cap), rotates
    to the next key in the pool automatically and keeps going. Only raises
    RateLimitError once every available key has been tried and failed."""
    img = prep_image_for_upload(pil_img)

    keys_tried = 0
    last_error = None
    while keys_tried < len(API_KEYS):
        current_model = get_model()
        for attempt in range(max_retries_per_key):
            try:
                resp = current_model.generate_content([EXTRACTION_PROMPT, img])
                raw = resp.text.strip()
                if raw.startswith("```"):
                    raw = raw.strip("`")
                    raw = raw.replace("json\n", "", 1) if raw.startswith("json\n") else raw
                data = json.loads(raw)
                data["_source_file"] = uf_name
                data["_uploaded_at"] = datetime.now().strftime("%Y-%m-%d %H:%M")
                data["_customer_tag"] = customer_tag
                data["_math_issues"] = verify_invoice_math(data)
                return data
            except Exception as e:
                if _is_rate_limit_error(e):
                    last_error = e
                    if attempt < max_retries_per_key - 1:
                        time.sleep(2 ** (attempt + 1))
                        continue
                    break
                raise

        keys_tried += 1
        if not rotate_to_next_key():
            break

    raise RateLimitError(
        f"All {len(API_KEYS)} configured Gemini key(s) are currently rate-limited "
        "(free tier: about 5 requests/minute and 20/day per key). Wait a few minutes, "
        "or add another GEMINI_API_KEY_N in Secrets for automatic fallback."
    ) from last_error



# ============================================================================
# ADVANCED FEATURES: GSTIN checksum validation, AI insights, anomaly detection, PDF export
# ============================================================================

GSTIN_ALPHABET = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"


def validate_gstin(gstin):
    """Validates an Indian GSTIN using its real check-digit algorithm (mod-36),
    the same family of checksum as a credit card's Luhn check but base-36.
    Returns (is_valid: bool, reason: str)."""
    if not gstin:
        return None, "Not provided"
    g = gstin.strip().upper().replace(" ", "")
    if len(g) != 15:
        return False, f"GSTIN must be 15 characters (found {len(g)})"
    if not all(c in GSTIN_ALPHABET for c in g):
        return False, "Contains invalid characters"
    try:
        total = 0
        for i, ch in enumerate(g[:14]):
            digit = GSTIN_ALPHABET.index(ch)
            factor = 2 if i % 2 == 1 else 1
            prod = digit * factor
            prod = (prod // 36) + (prod % 36)
            total += prod
        check_digit = (36 - (total % 36)) % 36
        expected = GSTIN_ALPHABET[check_digit]
        if expected == g[14]:
            return True, "Valid checksum"
        return False, f"Checksum mismatch (expected '{expected}', found '{g[14]}')"
    except Exception:
        return False, "Could not validate"


def detect_anomalies(invoices):
    """Flags line items priced noticeably above a vendor's own historical average
    for that exact product — real overcharge detection, not just a display table.
    Returns a list of anomaly dicts."""
    # build vendor+item -> list of (rate, source) history
    history = {}
    for inv in invoices:
        vendor = inv.get("vendor_name", "") or "Unknown Vendor"
        for item in inv.get("line_items", []):
            particulars = (item.get("particulars", "") or "").strip().lower()
            if not particulars:
                continue
            rate, ok = to_decimal(item.get("rate"))
            if ok and rate is not None:
                history.setdefault((vendor, particulars), []).append(
                    (rate, inv.get("bill_no", ""), inv.get("_source_file", ""))
                )

    anomalies = []
    THRESHOLD_PCT = Decimal("20")  # flag if >20% above that item's own historical average
    for (vendor, particulars), records in history.items():
        if len(records) < 2:
            continue  # need at least 2 data points to have a "usual" rate
        rates = [r[0] for r in records]
        avg_rate = sum(rates) / len(rates)
        if avg_rate == 0:
            continue
        for rate, bill_no, source_file in records:
            pct_over = ((rate - avg_rate) / avg_rate) * 100
            if pct_over > THRESHOLD_PCT:
                anomalies.append({
                    "vendor": vendor, "particulars": particulars.title(),
                    "bill_no": bill_no, "file": source_file,
                    "rate": float(rate), "avg_rate": float(avg_rate.quantize(Decimal("0.01"))),
                    "pct_over": float(pct_over.quantize(Decimal("0.1"))),
                })
    return sorted(anomalies, key=lambda a: -a["pct_over"])


def generate_ai_insights(invoices):
    """Sends aggregated (non-sensitive, numbers-only) stats to Gemini and asks for a
    short plain-English summary — the 'this reads like a junior accountant's note'
    feature. Cheap: one small text call, not a vision call."""
    if not invoices:
        return "No invoices yet — add some to generate insights."

    by_vendor = {}
    for inv in invoices:
        v = inv.get("vendor_name", "") or "Unknown Vendor"
        by_vendor.setdefault(v, []).append(inv)

    stats_lines = []
    total_spend = Decimal("0")
    total_gst = Decimal("0")
    for vendor, invs in by_vendor.items():
        spend = sum((d(i.get("grand_total")) for i in invs), Decimal("0"))
        gst = sum((d(i.get("cgst_amt")) + d(i.get("sgst_amt")) + d(i.get("igst_amt")) for i in invs), Decimal("0"))
        total_spend += spend
        total_gst += gst
        stats_lines.append(f"- {vendor}: {len(invs)} invoice(s), spend ₹{spend}, GST paid ₹{gst}")

    prompt = f"""You are a sharp junior accountant writing a 3-4 sentence plain-English note for a
small business owner reviewing their vendor invoices. Be specific and use the real numbers given.
Mention the largest vendor by spend, the total GST paid, and one observation worth their attention
(e.g. a vendor with unusually many invoices, or a concentration risk). No markdown, no bullet points,
just natural prose, like a short email note. Do not invent numbers not given below.

Total spend across all vendors: ₹{total_spend}
Total GST paid: ₹{total_gst}
Per-vendor breakdown:
{chr(10).join(stats_lines)}
"""
    resp = get_model().generate_content(prompt)

    return resp.text.strip()


class VendorStatementPDF(FPDF):
    def header(self):
        self.set_font("Helvetica", "B", 16)
        self.set_text_color(30, 30, 30)
        self.cell(0, 10, "GST Vendor Statement", ln=True)
        self.set_draw_color(180, 140, 80)
        self.set_line_width(0.6)
        self.line(10, 20, 200, 20)
        self.ln(6)


def generate_vendor_pdf(vendor_name, vendor_gstin, invoices):
    """One-page professional PDF statement for a single vendor — totals, tax
    breakdown, and every invoice, letterhead-style."""
    pdf = VendorStatementPDF(format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(50, 50, 50)

    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, vendor_name, ln=True)
    pdf.set_font("Helvetica", "", 10)
    is_valid, reason = validate_gstin(vendor_gstin)
    gstin_note = f"GSTIN: {vendor_gstin}" if vendor_gstin else "GSTIN: not on file"
    if is_valid is False:
        gstin_note += f"  [WARNING: {reason}]"
    pdf.cell(0, 6, gstin_note, ln=True)
    pdf.cell(0, 6, f"Statement generated: {datetime.now().strftime('%d %b %Y')}", ln=True)
    pdf.ln(4)

    total_spend = sum((d(i.get("grand_total")) for i in invoices), Decimal("0"))
    total_cgst = sum((d(i.get("cgst_amt")) for i in invoices), Decimal("0"))
    total_sgst = sum((d(i.get("sgst_amt")) for i in invoices), Decimal("0"))
    total_igst = sum((d(i.get("igst_amt")) for i in invoices), Decimal("0"))

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Summary", ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Invoices: {len(invoices)}     Total spend: Rs. {total_spend:,.2f}", ln=True)
    pdf.cell(0, 6, f"CGST paid: Rs. {total_cgst:,.2f}   SGST paid: Rs. {total_sgst:,.2f}   IGST paid: Rs. {total_igst:,.2f}", ln=True)
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Invoices", ln=True)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(230, 230, 230)
    col_widths = [30, 30, 40, 45, 45]
    headers = ["Bill No.", "Date", "Subtotal", "Tax", "Grand Total"]
    for w, h in zip(col_widths, headers):
        pdf.cell(w, 7, h, border=1, fill=True)
    pdf.ln()
    pdf.set_font("Helvetica", "", 9)
    for inv in invoices:
        tax = d(inv.get("cgst_amt")) + d(inv.get("sgst_amt")) + d(inv.get("igst_amt"))
        row = [
            str(inv.get("bill_no", ""))[:14],
            str(inv.get("date", ""))[:14],
            f"Rs. {float(d(inv.get('subtotal'))):,.2f}",
            f"Rs. {float(tax):,.2f}",
            f"Rs. {float(d(inv.get('grand_total'))):,.2f}",
        ]
        for w, val in zip(col_widths, row):
            pdf.cell(w, 6, val, border=1)
        pdf.ln()

    return bytes(pdf.output())
# This is what makes an admin view possible: everyone's uploads land in one shared
# file instead of each browser session having its own private, invisible copy.
# Caveat: on Streamlit Community Cloud, this file lives on ephemeral disk — it
# survives normal usage and app sleep/wake, but a fresh redeploy (new code push)
# wipes it. For real production durability, this should move to Supabase Postgres
# later; this gets you a genuinely working two-interface app at zero cost today.
import sqlite3

DB_PATH = "gst_invoices.db"


def init_db():
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_tag TEXT,
            vendor_name TEXT,
            created_at TEXT,
            data_json TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    conn.commit()
    conn.close()


def get_active_key_index():
    """Shared across every user of the app — quota exhaustion happens at Google's
    account level, not per browser session, so rotation state must be global too."""
    conn = sqlite3.connect(DB_PATH)
    row = conn.execute("SELECT value FROM app_settings WHERE key = 'active_gemini_key_index'").fetchone()
    conn.close()
    return int(row[0]) if row else 0


def set_active_key_index(idx):
    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        "INSERT INTO app_settings (key, value) VALUES ('active_gemini_key_index', ?) "
        "ON CONFLICT(key) DO UPDATE SET value = excluded.value",
        (str(idx),)
    )
    conn.commit()
    conn.close()


def save_invoice_to_db(inv):
    conn = sqlite3.connect(DB_PATH)
    conn.execute(
        "INSERT INTO invoices (customer_tag, vendor_name, created_at, data_json) VALUES (?, ?, ?, ?)",
        (inv.get("_customer_tag", "unspecified"), inv.get("vendor_name", ""),
         inv.get("_uploaded_at", ""), json.dumps(inv))
    )
    conn.commit()
    conn.close()


def load_invoices_from_db(customer_tag=None):
    conn = sqlite3.connect(DB_PATH)
    if customer_tag:
        rows = conn.execute(
            "SELECT data_json FROM invoices WHERE customer_tag = ? ORDER BY id", (customer_tag,)
        ).fetchall()
    else:
        rows = conn.execute("SELECT data_json FROM invoices ORDER BY id").fetchall()
    conn.close()
    return [json.loads(r[0]) for r in rows]


def delete_all_for_customer(customer_tag):
    conn = sqlite3.connect(DB_PATH)
    conn.execute("DELETE FROM invoices WHERE customer_tag = ?", (customer_tag,))
    conn.commit()
    conn.close()


init_db()

# ---------- SIDEBAR: IDENTITY + MODE ----------
st.sidebar.title("🗒️ The Ledger")

if len(API_KEYS) > 1:
    active_idx = get_active_key_index()
    st.sidebar.caption(f"🔑 API key {active_idx + 1} of {len(API_KEYS)} active")

ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD", "")

if "is_admin" not in st.session_state:
    st.session_state.is_admin = False

with st.sidebar.expander("🔑 Admin login", expanded=False):
    if st.session_state.is_admin:
        st.markdown(stamp("Admin", "admin"), unsafe_allow_html=True)
        if st.button("Log out of admin view"):
            st.session_state.is_admin = False
            st.rerun()
    else:
        pw = st.text_input("Admin password", type="password", key="admin_pw_input")
        if st.button("Log in"):
            if ADMIN_PASSWORD and pw == ADMIN_PASSWORD:
                st.session_state.is_admin = True
                st.rerun()
            else:
                st.error("Incorrect password.")

if st.session_state.is_admin:
    st.sidebar.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)
    st.sidebar.subheader("Admin view")
    all_customer_tags = sorted(set(
        r[0] for r in sqlite3.connect(DB_PATH).execute(
            "SELECT DISTINCT customer_tag FROM invoices"
        ).fetchall()
    ))
    admin_filter = st.sidebar.selectbox(
        "View data for:", ["All customers"] + all_customer_tags
    )
    customer_name_input = None  # not used in admin mode
else:
    st.sidebar.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)
    st.sidebar.markdown("Enter your business name below — this identifies which invoices are yours.")
    customer_name_input = st.sidebar.text_input("Your business / customer name", "").strip()

# ---------- MAIN: ADD INVOICES (upload photo OR type manually) ----------
render_hero()

if st.session_state.is_admin:
    st.title("The Ledger — Admin")
    st.markdown('<div class="ledger-masthead">Every account, one book · filtered by the sidebar</div>', unsafe_allow_html=True)
else:
    st.title("The Ledger")
    st.markdown('<div class="ledger-masthead">GST invoice intelligence, exact to the rupee</div>', unsafe_allow_html=True)

    if not customer_name_input:
        st.info("👈 Enter your business name in the sidebar first — this keeps your invoices separate from other customers'.")
        st.stop()

if not st.session_state.is_admin:
    section("01", "Capture")
    st.markdown('''<div class="workflow-strip"><span class="workflow-step"><b>01</b> Upload or type</span><span class="workflow-arrow">→</span><span class="workflow-step"><b>02</b> Verify the numbers</span><span class="workflow-arrow">→</span><span class="workflow-step"><b>03</b> Export your ledger</span></div>''', unsafe_allow_html=True)
    tab_upload, tab_manual = st.tabs(["📷 Upload Photo(s)", "⌨️ Enter Manually"])

    with tab_upload:
        uploaded_files = st.file_uploader(
            "Upload invoice photo(s)", type=["jpg", "jpeg", "png"], accept_multiple_files=True,
            key="photo_uploader"
        )

        if uploaded_files and st.button("Extract data from uploaded photos", type="primary"):
            progress = st.progress(0, text="Starting extraction...")
            total = len(uploaded_files)
            done = 0
            rate_limited = False

            # Serialize requests to prevent a batch upload from bursting through
            # Gemini's per-minute quota. extract_one() retries and rotates keys.
            for uf in uploaded_files:
                fname = uf.name
                try:
                    data = extract_one(fname, Image.open(uf), customer_name_input)
                    save_invoice_to_db(data)
                except RateLimitError as e:
                    st.error(f"⚠️ {fname}: {e}")
                    rate_limited = True
                    break
                except Exception as e:
                    st.warning(f"Could not parse {fname}: {e}")
                done += 1
                progress.progress(done / total, text=f"Processed {done}/{total}...")

            progress.progress(1.0, text="Done.")
            if rate_limited:
                n_keys = len(API_KEYS)
                if n_keys > 1:
                    st.info(f"All {n_keys} configured API keys hit their rate limit. Wait a few minutes "
                            "for the per-minute limit to clear, or wait for the daily reset.")
                else:
                    st.info("Hit Gemini's free-tier limit (5/min or 20/day). Wait a minute and retry, "
                            "or add a second GEMINI_API_KEY_2 in Secrets for automatic fallback.")
            else:
                st.success(f"Processed {total} invoice(s).")
            st.rerun()

    with tab_manual:
        st.caption("For invoices you'd rather type in directly — no photo needed, no extraction uncertainty at all "
                   "since you're entering the exact figures yourself.")

        if "manual_line_items" not in st.session_state:
            st.session_state.manual_line_items = pd.DataFrame(
                [{"particulars": "", "hsn_code": "", "qty": "", "rate": "", "amount": ""}]
            )

        with st.form("manual_entry_form", clear_on_submit=False):
            c1, c2, c3 = st.columns(3)
            m_vendor = c1.text_input("Vendor name")
            m_vendor_gstin = c2.text_input("Vendor GSTIN (optional)")
            m_bill_no = c3.text_input("Bill No.")
            c4, c5, c6 = st.columns(3)
            m_date = c4.text_input("Date (as written, e.g. 6/4/26)")
            m_customer = c5.text_input("Customer name (optional)")
            m_customer_gstin = c6.text_input("Customer GSTIN (optional)")

            st.markdown("**Line items** — add a row per product/service. Amount auto-fills as qty × rate but you can override it.")
            edited_items = st.data_editor(
                st.session_state.manual_line_items,
                num_rows="dynamic",
                width="stretch",
                column_config={
                    "particulars": st.column_config.TextColumn("Particulars"),
                    "hsn_code": st.column_config.TextColumn("HSN Code"),
                    "qty": st.column_config.TextColumn("Qty"),
                    "rate": st.column_config.TextColumn("Rate (₹)"),
                    "amount": st.column_config.TextColumn("Amount (₹) — leave blank to auto-calc"),
                },
                key="manual_items_editor",
            )

            st.markdown("**Tax**")
            t1, t2, t3, t4, t5 = st.columns(5)
            m_subtotal = t1.text_input("Subtotal (₹)", "")
            m_cgst_pct = t2.text_input("CGST %", "")
            m_sgst_pct = t3.text_input("SGST %", "")
            m_igst_pct = t4.text_input("IGST %", "")
            m_grand_total = t5.text_input("Grand Total (₹)")

            submitted = st.form_submit_button("Add this invoice", type="primary")

            if submitted:
                if not m_vendor or not m_grand_total:
                    st.error("Vendor name and Grand Total are required at minimum.")
                else:
                    line_items = []
                    for _, row in edited_items.iterrows():
                        if not str(row.get("particulars", "")).strip():
                            continue
                        qty, _ = extract_leading_number(row.get("qty"))
                        rate, _ = to_decimal(row.get("rate"))
                        amt_raw = str(row.get("amount", "")).strip()
                        if amt_raw:
                            amount_str = amt_raw
                        elif qty is not None and rate is not None:
                            amount_str = str((qty * rate).quantize(Decimal("0.01")))
                        else:
                            amount_str = ""
                        line_items.append({
                            "particulars": row.get("particulars", ""),
                            "hsn_code": row.get("hsn_code", ""),
                            "qty": str(row.get("qty", "")),
                            "rate": str(row.get("rate", "")),
                            "amount": amount_str,
                        })

                    subtotal_val = m_subtotal.strip()
                    if not subtotal_val and line_items:
                        auto_sub = sum((d(li["amount"]) for li in line_items), Decimal("0"))
                        subtotal_val = str(auto_sub)

                    cgst_amt, sgst_amt, igst_amt = "", "", ""
                    sub_dec, sub_ok = to_decimal(subtotal_val)
                    if sub_ok and sub_dec is not None:
                        if m_cgst_pct.strip():
                            p, ok = to_decimal(m_cgst_pct)
                            if ok and p is not None:
                                cgst_amt = str((sub_dec * p / 100).quantize(Decimal("0.01")))
                        if m_sgst_pct.strip():
                            p, ok = to_decimal(m_sgst_pct)
                            if ok and p is not None:
                                sgst_amt = str((sub_dec * p / 100).quantize(Decimal("0.01")))
                        if m_igst_pct.strip():
                            p, ok = to_decimal(m_igst_pct)
                            if ok and p is not None:
                                igst_amt = str((sub_dec * p / 100).quantize(Decimal("0.01")))

                    data = {
                        "vendor_name": m_vendor,
                        "vendor_gstin": m_vendor_gstin,
                        "bill_no": m_bill_no,
                        "date": m_date,
                        "customer_name": m_customer,
                        "customer_gstin": m_customer_gstin,
                        "line_items": line_items,
                        "subtotal": subtotal_val,
                        "cgst_pct": m_cgst_pct, "cgst_amt": cgst_amt,
                        "sgst_pct": m_sgst_pct, "sgst_amt": sgst_amt,
                        "igst_pct": m_igst_pct, "igst_amt": igst_amt,
                        "grand_total": m_grand_total,
                        "uncertain_money_fields": [],  # you typed it yourself — nothing to flag as uncertain
                        "extraction_confidence": "high",
                        "_source_file": f"Manual entry — {m_bill_no or 'no bill no.'}",
                        "_uploaded_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                        "_customer_tag": customer_name_input,
                    }
                    data["_math_issues"] = verify_invoice_math(data)  # still cross-checked, in case of a typo
                    save_invoice_to_db(data)
                    st.session_state.manual_line_items = pd.DataFrame(
                        [{"particulars": "", "hsn_code": "", "qty": "", "rate": "", "amount": ""}]
                    )
                    st.success(f"Added invoice from {m_vendor}. Scroll down to see it in the report.")
                    st.rerun()

st.divider()

# ---------- LOAD DATA FOR DISPLAY (from shared DB, scoped by mode) ----------
if st.session_state.is_admin:
    if admin_filter == "All customers":
        st.session_state.invoices = load_invoices_from_db()
    else:
        st.session_state.invoices = load_invoices_from_db(admin_filter)
else:
    st.session_state.invoices = load_invoices_from_db(customer_name_input)

# ---------- RESULTS TABLE ----------
if st.session_state.invoices:
    section("02", "Ledger")
    st.subheader("Extracted invoices")

    rows = []
    for inv in st.session_state.invoices:
        issues = inv.get("_math_issues", [])
        uncertain = inv.get("uncertain_money_fields", [])
        gstin_valid, gstin_reason = validate_gstin(inv.get("vendor_gstin", ""))
        gstin_label = "—" if gstin_valid is None else ("Valid" if gstin_valid else "Invalid")
        rows.append({
            "File": inv.get("_source_file", ""),
            "Customer tag": inv.get("_customer_tag", ""),
            "Vendor": inv.get("vendor_name", ""),
            "Vendor GSTIN": inv.get("vendor_gstin", "") or "—",
            "GSTIN Check": gstin_label,
            "Bill No.": inv.get("bill_no", ""),
            "Date": inv.get("date", ""),
            "Subtotal": float(d(inv.get("subtotal"))),
            "CGST": float(d(inv.get("cgst_amt"))),
            "SGST": float(d(inv.get("sgst_amt"))),
            "IGST": float(d(inv.get("igst_amt"))),
            "Grand Total": float(d(inv.get("grand_total"))),
            "Confidence": inv.get("extraction_confidence", ""),
            "Numbers OK?": "Needs review" if (issues or uncertain) else "Accurate",
        })
    df = pd.DataFrame(rows)

    # style the status column with the ledger stamp badge inline with the paper-panel table
    def _stamp_style(val):
        if val in ("Accurate", "Valid"):
            return "color: #3F8F5F; font-weight: 600; font-family: 'IBM Plex Mono', monospace;"
        if val in ("Needs review", "Invalid"):
            return "color: #C1553A; font-weight: 600; font-family: 'IBM Plex Mono', monospace;"
        return ""

    styled_df = df.style.map(_stamp_style, subset=["Numbers OK?", "GSTIN Check"]).format(
        {"Subtotal": "₹{:.2f}", "CGST": "₹{:.2f}", "SGST": "₹{:.2f}", "IGST": "₹{:.2f}", "Grand Total": "₹{:.2f}"}
    )
    st.dataframe(styled_df, width="stretch")

    n_invalid_gstin = sum(1 for r in rows if r["GSTIN Check"] == "Invalid")
    if n_invalid_gstin:
        st.markdown(stamp(f"{n_invalid_gstin} vendor GSTIN(s) fail checksum validation", "flagged"), unsafe_allow_html=True)
        st.caption("A GSTIN that fails its check-digit is either mistyped on the invoice or the invoice itself is questionable — worth confirming with the vendor.")

    # ---------- FLAGGED INVOICES — only for genuine money-accuracy problems.
    # Text fields (customer name/GSTIN, date formatting) never trigger this anymore —
    # only a real math mismatch or a money digit the model wasn't confident reading. ----------
    any_issues = any(inv.get("_math_issues") for inv in st.session_state.invoices)
    any_uncertain = any(inv.get("uncertain_money_fields") for inv in st.session_state.invoices)
    if any_issues or any_uncertain:
        n_flagged = sum(1 for inv in st.session_state.invoices
                         if inv.get("_math_issues") or inv.get("uncertain_money_fields"))
        st.markdown(
            stamp(f"{n_flagged} of {len(st.session_state.invoices)} need review", "flagged"),
            unsafe_allow_html=True
        )
        st.caption("Everything else is accurate as extracted — nothing here is auto-corrected, only flagged.")
        for inv in st.session_state.invoices:
            issues = inv.get("_math_issues", [])
            uncertain = inv.get("uncertain_money_fields", [])
            if issues or uncertain:
                with st.expander(f"{inv.get('_source_file','')} — {inv.get('vendor_name','')} "
                                  f"(Bill {inv.get('bill_no','')})", expanded=False):
                    if issues:
                        st.markdown("**Math discrepancies (recomputed independently, exact decimal math):**")
                        for iss in issues:
                            st.markdown(f"- {iss}")
                    if uncertain:
                        st.markdown("**Money figures the model wasn't fully confident reading:**")
                        st.markdown(f"- {', '.join(uncertain)}")
                    st.caption("Open the original photo and correct these manually before relying on this invoice's totals.")
    else:
        st.markdown(stamp("All figures verified", "verified"), unsafe_allow_html=True)

    # ---------- LINE ITEM DETAIL ----------
    with st.expander("View line-item detail per invoice"):
        for inv in st.session_state.invoices:
            st.markdown(f"**{inv.get('vendor_name','')} — Bill {inv.get('bill_no','')} ({inv.get('date','')})**")
            items = inv.get("line_items", [])
            if items:
                st.table(pd.DataFrame(items))
            else:
                st.caption("No line items extracted.")

    # ---------- ANALYTICS ----------
    section("03", "Analytics")
    st.subheader("Analytics")

    # exact Decimal totals for headline metrics — summed before any float conversion
    total_spend = sum((d(inv.get("grand_total")) for inv in st.session_state.invoices), Decimal("0"))
    total_cgst = sum((d(inv.get("cgst_amt")) for inv in st.session_state.invoices), Decimal("0"))
    total_sgst = sum((d(inv.get("sgst_amt")) for inv in st.session_state.invoices), Decimal("0"))
    total_igst = sum((d(inv.get("igst_amt")) for inv in st.session_state.invoices), Decimal("0"))
    total_gst = total_cgst + total_sgst + total_igst

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total invoices", len(df))
    col2.metric("Total spend (₹)", f"{total_spend:,.2f}")
    col3.metric("Total GST paid (₹)", f"{total_gst:,.2f}")
    col4.metric("Effective GST rate", f"{(total_gst / total_spend * 100):.2f}%" if total_spend else "—")

    # ---- AI-GENERATED INSIGHTS ----
    st.markdown('<div class="glass-panel">', unsafe_allow_html=True)
    st.markdown("**🧠 AI Insights**")
    insights_key = f"insights_{len(st.session_state.invoices)}"  # regenerate if invoice count changes
    if st.button("Generate insights", key="gen_insights_btn"):
        with st.spinner("Reading through the ledger..."):
            st.session_state[insights_key] = generate_ai_insights(st.session_state.invoices)
    if insights_key in st.session_state:
        st.markdown(f"*{st.session_state[insights_key]}*")
    else:
        st.caption("Click to have the AI read your data and write a short plain-English summary.")
    st.markdown('</div>', unsafe_allow_html=True)
    st.write("")

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(
        ["By Vendor", "By HSN Code", "By Item", "Monthly Trend", "Tax Breakdown", "⚠️ Anomalies"]
    )

    with tab1:
        vendor_summary = df.groupby("Vendor")["Grand Total"].sum().sort_values(ascending=False)
        st.bar_chart(vendor_summary)
        st.dataframe(vendor_summary.reset_index().rename(columns={"Grand Total": "Total Spend (₹)"}),
                     width="stretch")

    with tab2:
        hsn_rows = []
        for inv in st.session_state.invoices:
            for item in inv.get("line_items", []):
                hsn_rows.append({
                    "HSN Code": item.get("hsn_code", "") or "unspecified",
                    "Amount": float(d(item.get("amount"))),
                })
        if hsn_rows:
            hsn_df = pd.DataFrame(hsn_rows)
            hsn_summary = hsn_df.groupby("HSN Code")["Amount"].sum().sort_values(ascending=False)
            st.bar_chart(hsn_summary)
            st.dataframe(hsn_summary.reset_index().rename(columns={"Amount": "Total Amount (₹)"}),
                         width="stretch")
        else:
            st.caption("No line items to break down yet.")

    with tab3:
        item_rows = []
        for inv in st.session_state.invoices:
            for item in inv.get("line_items", []):
                qty = d(item.get("qty"))
                rate = d(item.get("rate"))
                item_rows.append({
                    "Particulars": item.get("particulars", ""),
                    "Vendor": inv.get("vendor_name", ""),
                    "Qty": float(qty),
                    "Rate": float(rate),
                    "Amount": float(d(item.get("amount"))),
                })
        if item_rows:
            item_df = pd.DataFrame(item_rows)
            st.caption("Same item, different rate across vendors/dates — useful for spotting overcharging.")
            st.dataframe(
                item_df.groupby(["Particulars", "Vendor"]).agg(
                    Times_Billed=("Amount", "count"),
                    Avg_Rate=("Rate", "mean"),
                    Total_Amount=("Amount", "sum"),
                ).reset_index().sort_values("Total_Amount", ascending=False),
                width="stretch",
            )
        else:
            st.caption("No line items to break down yet.")

    with tab4:
        trend_rows = []
        for inv in st.session_state.invoices:
            trend_rows.append({"Date": inv.get("date", "") or "unspecified",
                                "Grand Total": float(d(inv.get("grand_total")))})
        trend_df = pd.DataFrame(trend_rows)
        monthly = trend_df.groupby("Date")["Grand Total"].sum()
        st.line_chart(monthly)
        st.caption("Grouped by the date exactly as it appears on each invoice — normalize date formats "
                    "manually if vendors write dates inconsistently.")

    with tab5:
        tax_df = pd.DataFrame({
            "Tax type": ["CGST", "SGST", "IGST"],
            "Amount (₹)": [float(total_cgst), float(total_sgst), float(total_igst)],
        })
        st.bar_chart(tax_df.set_index("Tax type"))
        st.dataframe(tax_df, width="stretch")

    with tab6:
        anomalies = detect_anomalies(st.session_state.invoices)
        if anomalies:
            st.markdown(stamp(f"{len(anomalies)} rate(s) above vendor's own historical average", "flagged"),
                        unsafe_allow_html=True)
            st.caption("Flagged when a line item's rate is more than 20% above that same vendor's own "
                       "historical average rate for the identical product — a real overcharge signal, "
                       "not just a display of the numbers.")
            anom_df = pd.DataFrame(anomalies).rename(columns={
                "vendor": "Vendor", "particulars": "Item", "bill_no": "Bill No.", "file": "File",
                "rate": "Billed Rate (₹)", "avg_rate": "Vendor's Usual Rate (₹)", "pct_over": "% Above Usual",
            })
            st.dataframe(
                anom_df.style.map(lambda v: "color: #C1553A; font-weight: 600;", subset=["% Above Usual"])
                       .format({"Billed Rate (₹)": "₹{:.2f}", "Vendor's Usual Rate (₹)": "₹{:.2f}", "% Above Usual": "+{:.1f}%"}),
                width="stretch"
            )
        else:
            st.markdown(stamp("No overcharge patterns detected", "verified"), unsafe_allow_html=True)
            st.caption("Needs at least 2 invoices with the same item from the same vendor to establish a baseline rate.")

    low_conf = df[df["Confidence"] == "low"]
    if not low_conf.empty:
        st.warning(f"{len(low_conf)} invoice(s) had low extraction confidence — worth a manual double-check "
                    "against the original photo.")

    # ---------- EXPORT ----------
    section("04", "Export")
    st.subheader("Export")

    import openpyxl
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    def safe_sheet_name(name, used):
        clean = "".join(c for c in (name or "Unknown Vendor") if c not in r'[]:*?/\\')[:28].strip() or "Unknown Vendor"
        candidate, i = clean, 1
        while candidate in used:
            i += 1
            candidate = f"{clean[:25]} ({i})"
        used.add(candidate)
        return candidate

    HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    FLAG_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    def style_header(ws, row_num, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

    def autosize(ws, ncols, min_w=10, max_w=45):
        for c in range(1, ncols + 1):
            letter = get_column_letter(c)
            longest = max(
                [len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1)] + [min_w]
            )
            ws.column_dimensions[letter].width = min(longest + 2, max_w)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names = set()

    # group invoices by vendor
    by_vendor = {}
    for inv in st.session_state.invoices:
        vname = inv.get("vendor_name") or "Unknown Vendor"
        by_vendor.setdefault(vname, []).append(inv)

    # ---- OVERVIEW SHEET ----
    ov = wb.create_sheet("Overview")
    headers = ["Vendor", "Invoices", "Total Spend (₹)", "Total CGST (₹)", "Total SGST (₹)",
               "Total IGST (₹)", "Total GST Paid (₹)", "Needs Review?"]
    ov.append(headers)
    style_header(ov, 1, len(headers))

    vendor_stats = []
    for vname, invs in by_vendor.items():
        spend = sum((d(i.get("grand_total")) for i in invs), Decimal("0"))
        cgst = sum((d(i.get("cgst_amt")) for i in invs), Decimal("0"))
        sgst = sum((d(i.get("sgst_amt")) for i in invs), Decimal("0"))
        igst = sum((d(i.get("igst_amt")) for i in invs), Decimal("0"))
        needs_review = any(i.get("_math_issues") for i in invs)
        vendor_stats.append((vname, len(invs), spend, cgst, sgst, igst, cgst + sgst + igst, needs_review))

    for row in vendor_stats:
        vname, n, spend, cgst, sgst, igst, gst_total, needs_review = row
        ov.append([vname, n, float(spend), float(cgst), float(sgst), float(igst),
                   float(gst_total), "⚠️ Yes" if needs_review else "OK"])
        if needs_review:
            for c in range(1, len(headers) + 1):
                ov.cell(row=ov.max_row, column=c).fill = FLAG_FILL

    grand_row = ov.max_row + 2
    ov.cell(row=grand_row, column=1, value="TOTAL").font = Font(bold=True)
    ov.cell(row=grand_row, column=3, value=sum(v[2] for v in vendor_stats)).font = Font(bold=True)
    ov.cell(row=grand_row, column=7, value=sum(v[6] for v in vendor_stats)).font = Font(bold=True)
    autosize(ov, len(headers))

    if vendor_stats:
        chart = BarChart()
        chart.title = "GST paid by vendor"
        chart.y_axis.title = "₹"
        n_vendors = len(vendor_stats)
        data = Reference(ov, min_col=7, min_row=1, max_row=1 + n_vendors)
        cats = Reference(ov, min_col=1, min_row=2, max_row=1 + n_vendors)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width, chart.height = 18, 10
        ov.add_chart(chart, f"J2")

        pie = PieChart()
        pie.title = "Total spend share by vendor"
        pdata = Reference(ov, min_col=3, min_row=1, max_row=1 + n_vendors)
        pie.add_data(pdata, titles_from_data=True)
        pie.set_categories(cats)
        pie.width, pie.height = 14, 10
        ov.add_chart(pie, f"J22")

    # ---- ONE SHEET PER VENDOR ----
    for vname, invs in by_vendor.items():
        sheet_name = safe_sheet_name(vname, used_names)
        ws = wb.create_sheet(sheet_name)

        ws.cell(row=1, column=1, value=vname).font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"GSTIN: {invs[0].get('vendor_gstin','') or '—'}")

        # Line-item level detail: product, qty, rate, amount, and the invoice-level GST that applies to it
        headers = ["Bill No.", "Date", "Particulars", "HSN Code", "Qty", "Rate (₹)", "Amount (₹)",
                   "CGST % (bill)", "SGST % (bill)", "Needs Review?"]
        start_row = 4
        ws.append([])  # row 3 blank
        ws.append(headers)
        style_header(ws, start_row + 1, len(headers))

        r = start_row + 2
        for inv in invs:
            flagged = bool(inv.get("_math_issues"))
            for item in inv.get("line_items", []):
                ws.append([
                    inv.get("bill_no", ""),
                    inv.get("date", ""),
                    item.get("particulars", ""),
                    item.get("hsn_code", ""),
                    str(item.get("qty", "")),
                    str(item.get("rate", "")),
                    float(d(item.get("amount"))),
                    str(inv.get("cgst_pct", "")),
                    str(inv.get("sgst_pct", "")),
                    "⚠️" if flagged else "OK",
                ])
                if flagged:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=r, column=c).fill = FLAG_FILL
                r += 1

        items_end_row = r - 1
        autosize(ws, len(headers))

        # Invoice-level summary block (subtotal / tax / grand total per bill) below the line items
        r += 2
        ws.cell(row=r, column=1, value="Invoice-level summary").font = Font(bold=True, size=12)
        r += 1
        sum_headers = ["Bill No.", "Date", "Subtotal (₹)", "CGST (₹)", "SGST (₹)", "IGST (₹)", "Grand Total (₹)"]
        for c, h in enumerate(sum_headers, start=1):
            ws.cell(row=r, column=c, value=h)
        style_header(ws, r, len(sum_headers))
        summary_start = r + 1
        r += 1
        for inv in invs:
            ws.append_row = None  # no-op, keeping structure clear
            ws.cell(row=r, column=1, value=inv.get("bill_no", ""))
            ws.cell(row=r, column=2, value=inv.get("date", ""))
            ws.cell(row=r, column=3, value=float(d(inv.get("subtotal"))))
            ws.cell(row=r, column=4, value=float(d(inv.get("cgst_amt"))))
            ws.cell(row=r, column=5, value=float(d(inv.get("sgst_amt"))))
            ws.cell(row=r, column=6, value=float(d(inv.get("igst_amt"))))
            ws.cell(row=r, column=7, value=float(d(inv.get("grand_total"))))
            r += 1
        summary_end = r - 1

        # Chart: spend per bill for this vendor
        if summary_end >= summary_start:
            vchart = BarChart()
            vchart.title = f"{vname} — spend per invoice"
            vchart.y_axis.title = "₹"
            data = Reference(ws, min_col=7, min_row=summary_start - 1, max_row=summary_end)
            cats = Reference(ws, min_col=1, min_row=summary_start, max_row=summary_end)
            vchart.add_data(data, titles_from_data=True)
            vchart.set_categories(cats)
            vchart.width, vchart.height = 16, 9
            ws.add_chart(vchart, f"L4")

    # ---- FLAGS SHEET (all vendors, needs-review items in one place) ----
    flags_rows = []
    for inv in st.session_state.invoices:
        for iss in inv.get("_math_issues", []):
            flags_rows.append((inv.get("vendor_name", ""), inv.get("_source_file", ""), "Math", iss))
        for u in inv.get("uncertain_money_fields", []):
            flags_rows.append((inv.get("vendor_name", ""), inv.get("_source_file", ""), "Uncertain money field", u))
    if flags_rows:
        fs = wb.create_sheet("Flags - Needs Review")
        fs.append(["Vendor", "File", "Type", "Issue"])
        style_header(fs, 1, 4)
        for row in flags_rows:
            fs.append(list(row))
        autosize(fs, 4)

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)

    st.download_button(
        "Download Excel report",
        data=excel_buffer.getvalue(),
        file_name=f"gst_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption("One sheet per vendor (products, rates, GST %, invoice totals + a spend chart), "
               "plus an Overview sheet with vendor-wise GST and spend charts.")

    # ---- PDF STATEMENTS: one professional one-pager per vendor ----
    st.markdown("**📄 PDF Statements** — a clean, one-page summary per vendor, ready to hand someone directly.")
    for vname, invs in by_vendor.items():
        pdf_bytes = generate_vendor_pdf(vname, invs[0].get("vendor_gstin", ""), invs)
        st.download_button(
            f"Download PDF — {vname}",
            data=pdf_bytes,
            file_name=f"{vname.replace(' ', '_')}_statement.pdf",
            mime="application/pdf",
            key=f"pdf_{vname}",
        )

    if not st.session_state.is_admin:
        if st.button("Clear my data"):
            delete_all_for_customer(customer_name_input)
            st.rerun()
    else:
        st.caption("Admins can't bulk-delete a customer's data from here — this protects against "
                   "accidentally wiping a paying customer's records. Use direct DB access if truly needed.")
else:
    st.info("No invoices processed yet. Upload photos above to get started.")
